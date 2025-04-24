from fastapi import FastAPI, File, UploadFile
from fastapi.responses import JSONResponse
from openpyxl import load_workbook
from PIL import Image
from io import BytesIO
import base64
import boto3
import os
from dotenv import load_dotenv
import uuid

app = FastAPI()

# Load environment variables from .env file
load_dotenv()

# Load AWS credentials and S3 config from environment variables
S3_AWS_ACCESS_KEY = os.getenv("S3_AWS_ACCESS_KEY")
S3_AWS_SECRET_KEY = os.getenv("S3_AWS_SECRET_KEY")
S3_BUCKET = os.getenv("S3_BUCKET")
S3_REGION = os.getenv("S3_REGION")

s3_client = boto3.client(
    "s3",
    aws_access_key_id=S3_AWS_ACCESS_KEY,
    aws_secret_access_key=S3_AWS_SECRET_KEY,
    region_name=S3_REGION
)

@app.post("/extract-images")
async def extract_images(file: UploadFile = File(...)):
    try:
        file_data = await file.read()
        wb = load_workbook(filename=BytesIO(file_data))
        
        result = {}

        for idx, sheet_name in enumerate(wb.sheetnames):
            
            print(f"Processing sheet: {sheet_name}")
            
            ws = wb[sheet_name]
            sheet_key = f"{idx}"
            extracted = []

            for img in ws._images:
                try:
                    anchor = img.anchor._from
                    row = anchor.row + 1  # Excel row (0-indexed to 1-indexed)
                    col = anchor.col + 1  # Excel column (0-indexed to 1-indexed)

                    img_bytes = BytesIO()
                    pil_img = Image.open(img.ref)
                    pil_img.save(img_bytes, format="PNG")
                    # Build base64 like frontend: "data:image/png;base64,...."

                    encoded_str = base64.b64encode(img_bytes.getvalue()).decode()

                    # Decode base64 and upload to S3
                    decoded_image = base64.b64decode(encoded_str)

                    # Generate a unique S3 key for the image
                    s3_key = f"{uuid.uuid4().hex}.png"


                    s3_client.put_object(
                        Bucket=S3_BUCKET,
                        Key=s3_key,
                        Body=decoded_image,
                        ACL='public-read',
                        ContentType="image/png"
                    )

                    image_url = f"https://{S3_BUCKET}.s3.{S3_REGION}.amazonaws.com/{s3_key}"

                    # Add the row and base64-encoded image to the extracted list
                    extracted.append({
                        "row": row,
                        "col": col,
                        "image_url": image_url,
                    })

                except Exception as e:
                    print(f"Error processing image on {sheet_name}: {e}")

            result[sheet_key] = extracted

        return JSONResponse(content=result)

    except Exception as e:
        return JSONResponse(
            status_code=500,
            content={"error": f"Failed to extract images: {str(e)}"}
        )

    # Load the Excel file from the uploaded bytes
    wb = load_workbook(filename=BytesIO(await file.read()))
    ws = wb.active  # Get the active worksheet

    extracted = []  # List to store the extracted image information

    # Iterate through the images in the worksheet
    for img in ws._images:
        try:
            # Get the anchor position of the image (this gives us the row and column)
            anchor = img.anchor._from
            row = anchor.row + 1  # Convert from 0-indexed to Excel row (1-indexed)

            # Open the image and convert it to a base64-encoded PNG
            img_bytes = BytesIO()
            pil_img = Image.open(img.ref)
            pil_img.save(img_bytes, format="PNG")  # Save as PNG in memory
            # Build base64 like frontend: "data:image/png;base64,...."

            encoded_str = base64.b64encode(img_bytes.getvalue()).decode()

            # Decode base64 and upload to S3
            decoded_image = base64.b64decode(encoded_str)

            # Generate a unique S3 key for the image
            s3_key = f"{uuid.uuid4().hex}.png"

            print('Uploading to S3:', s3_key)
            print('Bucket Name:', S3_BUCKET)
            print('Region:', S3_REGION)


            s3_client.put_object(
                Bucket=S3_BUCKET,
                Key=s3_key,
                Body=decoded_image,
                ACL='public-read',
                ContentType="image/png"
            )

            image_url = f"https://{S3_BUCKET}.s3.{S3_REGION}.amazonaws.com/{s3_key}"

            # Add the row and base64-encoded image to the extracted list
            extracted.append({
                "row": row,
                "image_url": image_url,
            })
        except Exception as e:
            # Handle any errors that occur while processing an image
            print(f"Error processing image: {e}")

    # Return the extracted data as JSON response
    return JSONResponse(content=extracted)
