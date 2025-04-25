# app/main.py
from fastapi import FastAPI, File, UploadFile
from fastapi.responses import JSONResponse
from openpyxl import load_workbook
from io import BytesIO
from app.utils import process_image
import os
from dotenv import load_dotenv
import boto3

load_dotenv()

S3_AWS_ACCESS_KEY = os.getenv("S3_AWS_ACCESS_KEY")
S3_AWS_SECRET_KEY = os.getenv("S3_AWS_SECRET_KEY")
S3_BUCKET = os.getenv("S3_BUCKET")
S3_REGION = os.getenv("S3_REGION")

s3_client = boto3.client(
    "s3",
    aws_access_key_id=S3_AWS_ACCESS_KEY,
    aws_secret_access_key=S3_AWS_SECRET_KEY,
    region_name=S3_REGION
)

app = FastAPI()

@app.post("/extract-images")
async def extract_images(file: UploadFile = File(...)):
    try:
        file_data = await file.read()
        wb = load_workbook(filename=BytesIO(file_data))

        result = {}

        for idx, sheet_name in enumerate(wb.sheetnames):
            print(f"Processing sheet: {sheet_name}")
            ws = wb[sheet_name]
            sheet_key = f"{idx}"
            extracted = []

            for img in ws._images:
                try:
                    processed = process_image(img, s3_client, S3_BUCKET, S3_REGION)
                    extracted.append(processed)
                except Exception as e:
                    print(f"Error processing image on {sheet_name}: {e}")

            result[sheet_key] = extracted

        return JSONResponse(content=result)

    except Exception as e:
        return JSONResponse(
            status_code=500,
            content={"error": f"Failed to extract images: {str(e)}"}
        )